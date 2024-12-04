import pathlib

import openpyxl
import pytest
from selenium.webdriver.chrome import webdriver


@pytest.fixture(scope="module")
def setup_driver():
    driver = webdriver.Chrome()
    driver.get('https://healthapp.yaksha.com/')
    driver.implicitly_wait(20)
    driver.maximize_window()
    yield driver
    driver.quit()


@pytest.fixture(scope="session")
def credentials():
    """
       Reads test credentials (username and password) from an Excel file.

       Returns:
           dict: A dictionary containing the test credentials.
       """
    file = pathlib.Path("./testData/Verification.xlsx")
    wb = openpyxl.load_workbook(file)
    sheet = wb["Credentials"]
    data = {}

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = row[0]
        value = row[1]
        data[key] = value
    return data


@pytest.fixture(scope="session")
def expected_data():
    """
        Reads expected test data for verification operations from an Excel file.

        Returns:
            dict: A dictionary containing the expected test data.
    """
    file = pathlib.Path("./testData/Verification.xlsx")
    wb = openpyxl.load_workbook(file)
    sheet = wb["Verification"]
    data = {}

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = row[0]
        value = row[1]
        data[key] = value

    return data